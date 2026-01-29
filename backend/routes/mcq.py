from flask import Blueprint, request, jsonify, send_file
from models import db, MCQ, MCQAttempt, Course, Student, StudentCourse, StudyMaterial, StaffCourse
from datetime import datetime
import openpyxl
from io import BytesIO

mcq_bp = Blueprint('mcq', __name__)

def serialize_mcq(m, include_answer=True):
    """Serialize MCQ - optionally hide correct answer for students"""
    data = {
        'id': m.id,
        'question_text': m.question_text,
        'option_a': m.option_a,
        'option_b': m.option_b,
        'option_c': m.option_c,
        'option_d': m.option_d,
        'marks': float(m.marks) if m.marks else 1.0,
        'course_id': m.course_id,
        'course_name': m.course.course_name if m.course else None,
        'staff_id': m.staff_id,
        'study_material_id': m.study_material_id,
        'created_at': m.created_at.isoformat() if m.created_at else None
    }
    if include_answer:
        data['correct_answer'] = m.correct_answer
    return data

# ----------------------
# Staff CRUD Operations
# ----------------------

@mcq_bp.route('/api/mcqs', methods=['GET'])
def get_all():
    """Get all MCQs with optional filters"""
    course_id = request.args.get('course_id')
    staff_id = request.args.get('staff_id')
    study_material_id = request.args.get('study_material_id')
    
    query = MCQ.query
    if course_id:
        query = query.filter_by(course_id=course_id)
    if staff_id:
        query = query.filter_by(staff_id=staff_id)
    if study_material_id:
        query = query.filter_by(study_material_id=study_material_id)
    
    mcqs = query.order_by(MCQ.created_at.desc()).all()
    return jsonify([serialize_mcq(m) for m in mcqs])

@mcq_bp.route('/api/mcqs/<int:mcq_id>', methods=['GET'])
def get_one(mcq_id):
    """Get single MCQ by ID"""
    mcq = MCQ.query.get_or_404(mcq_id)
    return jsonify(serialize_mcq(mcq))

@mcq_bp.route('/api/mcqs', methods=['POST'])
def create():
    """Create a new MCQ question"""
    data = request.get_json()
    
    mcq = MCQ(
        question_text=data['question_text'],
        option_a=data['option_a'],
        option_b=data['option_b'],
        option_c=data.get('option_c'),
        option_d=data.get('option_d'),
        correct_answer=data['correct_answer'].upper(),
        marks=data.get('marks', 1.0),
        course_id=data['course_id'],
        staff_id=data.get('staff_id'),
        study_material_id=data.get('study_material_id')
    )
    
    db.session.add(mcq)
    db.session.commit()
    return jsonify({'id': mcq.id, 'message': 'MCQ created successfully'}), 201

@mcq_bp.route('/api/mcqs/<int:mcq_id>', methods=['PUT'])
def update(mcq_id):
    """Update an existing MCQ"""
    mcq = MCQ.query.get_or_404(mcq_id)
    data = request.get_json()
    
    mcq.question_text = data.get('question_text', mcq.question_text)
    mcq.option_a = data.get('option_a', mcq.option_a)
    mcq.option_b = data.get('option_b', mcq.option_b)
    mcq.option_c = data.get('option_c', mcq.option_c)
    mcq.option_d = data.get('option_d', mcq.option_d)
    mcq.correct_answer = data.get('correct_answer', mcq.correct_answer).upper()
    mcq.marks = data.get('marks', mcq.marks)
    mcq.course_id = data.get('course_id', mcq.course_id)
    mcq.study_material_id = data.get('study_material_id', mcq.study_material_id)
    
    db.session.commit()
    return jsonify({'message': 'MCQ updated successfully'})

@mcq_bp.route('/api/mcqs/<int:mcq_id>', methods=['DELETE'])
def delete(mcq_id):
    """Delete an MCQ"""
    mcq = MCQ.query.get_or_404(mcq_id)
    db.session.delete(mcq)
    db.session.commit()
    return jsonify({'message': 'MCQ deleted successfully'})

@mcq_bp.route('/api/mcqs/import', methods=['POST'])
def import_mcqs():
    """Import MCQs from Excel file"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
        
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Invalid file type. Please upload .xlsx file'}), 400

    try:
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        
        # Expected headers
        headers = [cell.value for cell in sheet[1]]
        required_cols = ['question_text', 'option_a', 'option_b', 'correct_answer', 'marks']
        
        # Map headers to indices
        col_map = {h.lower(): i for i, h in enumerate(headers) if h}
        
        # Validate required columns
        missing = [col for col in required_cols if col not in col_map]
        if missing:
            return jsonify({'error': f'Missing required columns: {", ".join(missing)}'}), 400
            
        staff_id = request.form.get('staff_id')
        default_course_id = request.form.get('course_id')
        study_material_id = request.form.get('study_material_id')
        
        created_count = 0
        errors = []
        
        # Iterate rows (skip header)
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Helper to get value safely
                def get_val(col_name):
                    idx = col_map.get(col_name)
                    if idx is not None and idx < len(row):
                        return row[idx]
                    return None

                question_text = get_val('question_text')
                if not question_text: continue # Skip empty rows
                
                # Determine course - prioritize form selection
                course_id = default_course_id
                
                # Still allow course_code override IF form course not provided (legacy support)
                # But if form has course_id, we stick to it to avoid confusion
                if not course_id:
                    bg_course_code = get_val('course_code')
                    if bg_course_code:
                        course = Course.query.filter_by(course_code=str(bg_course_code).strip()).first()
                        if course:
                            course_id = course.id
                
                if not course_id:
                    errors.append(f"Row {i}: Course not specified and no default provided")
                    continue

                mcq = MCQ(
                    question_text=str(question_text),
                    option_a=str(get_val('option_a')),
                    option_b=str(get_val('option_b')),
                    option_c=str(get_val('option_c') or ''),
                    option_d=str(get_val('option_d') or ''),
                    correct_answer=str(get_val('correct_answer')).upper().strip(),
                    marks=float(get_val('marks') or 1.0),
                    course_id=course_id,
                    staff_id=staff_id,
                    study_material_id=study_material_id
                )
                db.session.add(mcq)
                created_count += 1
                
            except Exception as e:
                errors.append(f"Row {i}: {str(e)}")
        
        if created_count > 0:
            db.session.commit()
            
        return jsonify({
            'message': f'Successfully imported {created_count} MCQs',
            'errors': errors,
            'count': created_count
        })
        
    except Exception as e:
        return jsonify({'error': f'Failed to process file: {str(e)}'}), 500

@mcq_bp.route('/api/mcqs/template', methods=['GET'])
def download_template():
    """Download Excel template for MCQ import"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MCQ Template"
    
    # Headers
    headers = [
        'question_text', 'option_a', 'option_b', 'option_c', 'option_d', 
        'correct_answer', 'marks'
    ]
    ws.append(headers)
    
    # Sample row
    sample = [
        'What is the capital of France?', 'Berlin', 'Madrid', 'Paris', 'Rome', 
        'C', 1
    ]
    ws.append(sample)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output, 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='mcq_import_template.xlsx'
    )

# ----------------------
# Student Quiz Operations
# ----------------------

@mcq_bp.route('/api/courses/<int:course_id>/quiz', methods=['GET'])
def get_course_quiz(course_id):
    """Get all MCQs for a course (for students taking quiz - no answers)"""
    student_id = request.args.get('student_id')
    
    # Access control
    if student_id:
        student = Student.query.get_or_404(int(student_id))
        
        # Check explicit enrollment first
        enrollment = StudentCourse.query.filter_by(
            student_id=student.id, 
            course_id=course_id, 
            status='active'
        ).first()
        
        if not enrollment:
             return jsonify({'error': 'Access denied: You are not enrolled in this course'}), 403

    mcqs = MCQ.query.filter_by(course_id=course_id).all()
    
    # Get student's existing attempts for this course
    attempted_mcq_ids = []
    if student_id:
        attempts = MCQAttempt.query.filter_by(student_id=student_id).all()
        attempted_mcq_ids = [a.mcq_id for a in attempts]
    
    result = []
    for m in mcqs:
        mcq_data = serialize_mcq(m, include_answer=False)
        mcq_data['attempted'] = m.id in attempted_mcq_ids
        result.append(mcq_data)
    
    return jsonify({
        'course_id': course_id,
        'total_questions': len(mcqs),
        'attempted_count': len([m for m in mcqs if m.id in attempted_mcq_ids]),
        'questions': result
    })

@mcq_bp.route('/api/materials/<int:material_id>/quiz', methods=['GET'])
def get_material_quiz(material_id):
    """Get all MCQs for a specific study material (for students taking quiz)"""
    student_id = request.args.get('student_id')
    
    # 1. Access Control Logic
    if student_id:
        student = Student.query.get_or_404(int(student_id))
        material = StudyMaterial.query.get_or_404(material_id)
        
        # Determine course ID from material -> staff_course -> course
        staff_course = StaffCourse.query.get(material.staff_course_id)
        if not staff_course:
             return jsonify({'error': 'Invalid material context'}), 400
             
        course_id = staff_course.course_id
        
        # Check enrollment
        enrollment = StudentCourse.query.filter_by(
            student_id=student.id,
            course_id=course_id,
            status='active'
        ).first()
        
        if not enrollment:
             return jsonify({'error': 'Access denied: You are not enrolled in this course'}), 403

    mcqs = MCQ.query.filter_by(study_material_id=material_id).all()
    
    # Get student's existing attempts
    attempted_mcq_ids = []
    if student_id:
        attempts = MCQAttempt.query.filter_by(student_id=student_id).all()
        attempted_mcq_ids = [a.mcq_id for a in attempts]
    
    result = []
    for m in mcqs:
        mcq_data = serialize_mcq(m, include_answer=False)
        mcq_data['attempted'] = m.id in attempted_mcq_ids
        result.append(mcq_data)
    
    return jsonify({
        'material_id': material_id,
        'total_questions': len(mcqs),
        'attempted_count': len([m for m in mcqs if m.id in attempted_mcq_ids]),
        'questions': result
    })


@mcq_bp.route('/api/mcqs/<int:mcq_id>/attempt', methods=['POST'])
def submit_attempt(mcq_id):
    """Submit a student's answer for an MCQ"""
    mcq = MCQ.query.get_or_404(mcq_id)
    data = request.get_json()
    
    student_id = data['student_id']
    
    # 1. Access Control
    enrollment = StudentCourse.query.filter_by(
        student_id=student_id,
        course_id=mcq.course_id,
        status='active'
    ).first()
    
    if not enrollment:
         return jsonify({'error': 'Access denied: You are not enrolled in this course'}), 403

    selected_answer = data['selected_answer'].upper()
    
    # Check if already attempted
    existing = MCQAttempt.query.filter_by(
        student_id=student_id, 
        mcq_id=mcq_id
    ).first()
    
    if existing:
        return jsonify({'error': 'Already attempted this question'}), 400
    
    # Check if correct
    is_correct = (selected_answer == mcq.correct_answer)
    
    attempt = MCQAttempt(
        student_id=student_id,
        mcq_id=mcq_id,
        selected_answer=selected_answer,
        is_correct=is_correct
    )
    
    db.session.add(attempt)
    db.session.commit()
    
    return jsonify({
        'id': attempt.id,
        'is_correct': is_correct,
        'correct_answer': mcq.correct_answer,
        'marks_earned': float(mcq.marks) if is_correct else 0
    })

@mcq_bp.route('/api/student/<int:student_id>/quiz-results', methods=['GET'])
def get_student_results(student_id):
    """Get all quiz results for a student"""
    student = Student.query.get_or_404(student_id)
    
    # Get attempts grouped by course
    attempts = MCQAttempt.query.filter_by(student_id=student_id).all()
    
    # Organize by course
    course_stats = {}
    for attempt in attempts:
        mcq = attempt.mcq
        course_id = mcq.course_id
        
        if course_id not in course_stats:
            course = Course.query.get(course_id)
            course_stats[course_id] = {
                'course_id': course_id,
                'course_name': course.course_name if course else None,
                'course_code': course.course_code if course else None,
                'total_attempted': 0,
                'correct_count': 0,
                'total_marks': 0,
                'earned_marks': 0
            }
        
        course_stats[course_id]['total_attempted'] += 1
        course_stats[course_id]['total_marks'] += float(mcq.marks) if mcq.marks else 1
        if attempt.is_correct:
            course_stats[course_id]['correct_count'] += 1
            course_stats[course_id]['earned_marks'] += float(mcq.marks) if mcq.marks else 1
    
    # Calculate percentages
    for stats in course_stats.values():
        if stats['total_marks'] > 0:
            stats['percentage'] = round((stats['earned_marks'] / stats['total_marks']) * 100, 1)
        else:
            stats['percentage'] = 0
    
    return jsonify({
        'student_id': student_id,
        'student_name': student.full_name,
        'total_attempts': len(attempts),
        'total_correct': sum(1 for a in attempts if a.is_correct),
        'course_results': list(course_stats.values())
    })
